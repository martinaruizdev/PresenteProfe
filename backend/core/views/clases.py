from urllib.parse import quote

from django.conf import settings
from django.http import HttpResponse
from django.utils.timezone import localtime

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from core.models import Clase, Asistencia
from core.serializers import ClaseSerializer, AsistenciaSerializer
from api.permissions import IsDocente, IsAlumno
from core.qr import make_qr_token, verify_qr_token, QR_TTL_SECONDS

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ClaseViewSet(viewsets.ModelViewSet):
    serializer_class = ClaseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        p = getattr(user, "perfil", None)
        qs = Clase.objects.all()
        if p and p.rol == "DOCENTE":
            return qs.filter(materia__docente=user)
        return qs

    # ---------- QR ----------
    @action(
        detail=True,
        methods=["post"],
        permission_classes=[IsAuthenticated, IsDocente],
    )
    def qr(self, request, pk=None):
        clase = self.get_object()
        token = make_qr_token(clase.id)
        url = f"{settings.FRONTEND_BASE_URL}/checkin?clase_id={clase.id}&token={quote(token)}"
        return Response(
            {
                "token": token,
                "expires_in": QR_TTL_SECONDS,
                "checkin_url": url,
            }
        )

    # ---------- EXPORTAR ASISTENCIAS A EXCEL ----------
    @action(
        detail=True,
        methods=["get"],
        permission_classes=[IsAuthenticated, IsDocente],
        url_path="export-asistencias",
    )
    def export_asistencias(self, request, pk=None):
        """
        Exporta las asistencias en formato Excel
        """
        clase = self.get_object()

        # Obtener asistencias
        asistencias = (
            Asistencia.objects.filter(clase=clase)
            .select_related("alumno")
            .order_by("timestamp")
        )

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"

        # Encabezados
        headers = ["Nombre", "Email", "Método", "Fecha y Hora"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            # poner en negrita
            cell.font = cell.font.copy(bold=True)

        # Datos
        for row, asistencia in enumerate(asistencias, start=2):
            alumno = asistencia.alumno
            nombre = f"{alumno.first_name} {alumno.last_name}".strip() or alumno.username
            email = alumno.email or alumno.username

            # fecha/hora legible
            ts = asistencia.timestamp
            try:
                ts_local = localtime(ts)
            except Exception:
                ts_local = ts  # por si llega naive

            ws.cell(row=row, column=1, value=nombre)
            ws.cell(row=row, column=2, value=email)
            ws.cell(row=row, column=3, value=asistencia.metodo)
            ws.cell(row=row, column=4, value=ts_local.strftime("%d/%m/%Y %H:%M"))

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

        # Respuesta HTTP con archivo Excel
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename="Asistencias-Clase-{clase.id}.xlsx"'

        wb.save(response)
        return response

    # ---------- ESTADÍSTICAS ----------
    @action(
        detail=True,
        methods=["get"],
        permission_classes=[IsAuthenticated],
    )
    def estadisticas(self, request, pk=None):
        clase = self.get_object()
        total_alumnos = getattr(clase.materia, "cantidad_alumnos", 0)
        presentes = Asistencia.objects.filter(clase=clase).count()
        ausentes = max(total_alumnos - presentes, 0)
        return Response(
            {
                "clase_id": clase.id,
                "materia": clase.materia.nombre,
                "presentes": presentes,
                "ausentes": ausentes,
                "total_alumnos": total_alumnos,
            }
        )
